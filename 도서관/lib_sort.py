import os
from openpyxl import Workbook


path = "./Todo"
file_list = os.listdir(path)

from openpyxl import load_workbook

title = []
status = []

file_name = "./Todo/" + '05.28 명륜.xlsx'
wb= load_workbook(file_name, data_only=True)
ws = wb.active

t_range = 'A3:A238'
s_range = 'B3:B238'

title_range = ws[t_range]
for row in title_range:
    for cell in row:
        title.append(cell.value)

status_range = ws[s_range]
for row in status_range:
    for cell in row:
        status.append(cell.value)

Sort_title = []
Sort_status1 = []


count = 0
temp = []
i =0
p = 0
for tt in title:
    temp.append(tt)
    if i == 0:
        Sort_title.append(tt)
    if temp[i-1] != tt:
        Sort_status1.append(count)
        count = 0
        if status[i] == "대출가능" or status[i] == "관내대출가능":
            count = count + 1
        
        Sort_title.append(tt)
    else:
        if status[i] == "대출가능" or status[i] == "관내대출가능":
            count = count + 1
    i = i+1
B = [list(x) for x in zip(Sort_status1)]

wb2 = Workbook()
ws2 = wb2.active
for i in B:
    ws2.append(i)
wb2.save('exist.xlsx')
